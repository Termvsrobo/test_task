import datetime
from itertools import groupby
from pathlib import Path
from typing import BinaryIO, Dict, List, NoReturn, Union

import openpyxl
from django.db.models import Count, Q, Value

from .models import Order

_field_map = {
    'A': 'number',
    'B': 'state',
    'C': 'agreement',
    'D': 'status',
    'E': 'author',
    'F': 'filename',
    'G': 'created_at',
    'H': 'end_of_work_at',
    'I': 'duration',
    'J': 'fullname',
    'K': 'fullname_after_proccess',
    'L': 'materials_code',
    'M': 'same_materials',
    'N': 'bei',
    'O': 'ntd',
    'P': 'package_id',
}


def load_orders_from_excel(excel_file: Union[Path, BinaryIO], skip_first_line=True) -> NoReturn:
    """Процедура загружает заявки из excel в базу"""
    wb = openpyxl.load_workbook(excel_file)

    ws = wb["Data"]

    orders_from_excel_data = list()
    current_row = 0
    for row in ws.iter_rows():
        current_row += 1
        if skip_first_line and current_row == 1:
            continue
        _order_data = dict()
        for cell in row:
            _order_data[_field_map[cell.column_letter]] = (
                cell.value.strip() if cell.value is not None else cell.value
            )
        _order_data['created_at'] = datetime.datetime.strptime(_order_data['created_at'], '%d.%m.%Y %H:%M:%S')
        _order_data['end_of_work_at'] = datetime.datetime.strptime(
            _order_data['end_of_work_at'], '%d.%m.%Y %H:%M:%S'
        ) if _order_data['end_of_work_at'] else None
        orders_from_excel_data.append(Order(**_order_data))
    Order.objects.bulk_create(orders_from_excel_data)


def get_report_data(from_date: datetime.datetime, to_date: datetime.datetime) -> List[Dict]:
    """Функция возвращает аггрегированные данные по заявкам"""
    # Здесь можно спокойно выбрать необходимые данные одним таким запросом:
    # select
    #   'Загруженных заявок',
    #   COUNT(*) FILTER (WHERE created_at = '2023-08-22 09:43:50'),
    #   COUNT(*)
    # from orders_order

    # union all

    # select
    #   'Дубли',
    #   COUNT(*) FILTER (WHERE state like 'Дубликат%' and created_at = '2023-08-22 09:43:50'),
    #   COUNT(*) FILTER (WHERE state like 'Дубликат%')
    # from orders_order

    # union all

    # select
    #   'На создание',
    #   COUNT(*) FILTER (WHERE state like 'ДОБАВЛЕНИЕ%' and created_at = '2023-08-22 09:43:50'),
    #   COUNT(*) FILTER (WHERE state like 'ДОБАВЛЕНИЕ%')
    # from orders_order

    # union all

    # select
    #   'На расширение',
    #   COUNT(*) FILTER (WHERE state like 'РАСШИРЕНИЕ%' and created_at = '2023-08-22 09:43:50'),
    #   COUNT(*) FILTER (WHERE state like 'РАСШИРЕНИЕ%')
    # from orders_order

    # union all

    # select
    #   'Обработка завершена',
    #   COUNT(*) FILTER (WHERE end_of_work_at is not null and created_at = '2023-08-22 09:43:50'),
    #   COUNT(*) FILTER (WHERE end_of_work_at is not null)
    # from orders_order

    # union all

    # select
    #   'Возвращена на уточнение',
    #   COUNT(*) FILTER (WHERE status = 'Возвращена на уточнение' and created_at = '2023-08-22 09:43:50'),
    #   COUNT(*) FILTER (WHERE status = 'Возвращена на уточнение')
    # from orders_order

    # union all

    # select
    #   'Отправлена в обработку',
    #   COUNT(*) FILTER (WHERE status = 'Отправлена в обработку' and created_at = '2023-08-22 09:43:50'),
    #   COUNT(*) FILTER (WHERE status = 'Отправлена в обработку')
    # from orders_order

    # union all

    # select
    #   'Пакетов',
    #   COUNT(distinct package_id) FILTER (WHERE created_at = '2023-08-22 09:43:50'),
    #   COUNT(distinct package_id)
    # from orders_order

    # union all

    # select
    #   'Пользователей',
    #   COUNT(distinct author) FILTER (WHERE created_at = '2023-08-22 09:43:50'),
    #   COUNT(distinct author)
    # from orders_order
    # Но django в каждую аннотация добавляет GROUP BY, что ломает красивый результат.
    # Не нашел решения как убрать этот GROUP BY из запроса, не прибегая к нескольким запросам.
    # Это приводит к тому, что после получения данных приходится еще немного напильником доработать результат.

    result = []
    date_filter = Q(created_at__range=(from_date, to_date))
    qs = Order.objects.annotate(
        name_param=Value('Загруженных заявок')
    ).annotate(
        count_period=Count('id', filter=date_filter)
    ).annotate(
        count_all=Count('*')
    ).values('name_param', 'count_period', 'count_all')
    qs = qs.union(
        Order.objects.annotate(
            name_param=Value('Дубли')
        ).annotate(
            count_period=Count('state', filter=date_filter & Q(state__startswith='Дубликат'))
        ).annotate(
            count_all=Count('state', filter=Q(state__startswith='Дубликат'))
        ).values('name_param', 'count_period', 'count_all'),
        all=True
    )
    qs = qs.union(
        Order.objects.annotate(
            name_param=Value('На создание')
        ).annotate(
            count_period=Count('state', filter=date_filter & Q(state__startswith='ДОБАВЛЕНИЕ'))
        ).annotate(
            count_all=Count('state', filter=Q(state__startswith='ДОБАВЛЕНИЕ'))
        ).values('name_param', 'count_period', 'count_all'),
        all=True
    )
    qs = qs.union(
        Order.objects.annotate(
            name_param=Value('На расширение')
        ).annotate(
            count_period=Count('state', filter=date_filter & Q(state__startswith='РАСШИРЕНИЕ'))
        ).annotate(
            count_all=Count('state', filter=Q(state__startswith='РАСШИРЕНИЕ'))
        ).values('name_param', 'count_period', 'count_all'),
        all=True
    )
    qs = qs.union(
        Order.objects.annotate(
            name_param=Value('Обработка завершена')
        ).annotate(
            count_period=Count('status', filter=date_filter & Q(status=Order.StatusEnum.PROCESSED.value))
        ).annotate(
            count_all=Count('status', filter=Q(status=Order.StatusEnum.PROCESSED.value))
        ).values('name_param', 'count_period', 'count_all'),
        all=True
    )
    qs = qs.union(
        Order.objects.annotate(
            name_param=Value('Возвращена на уточнение')
        ).annotate(
            count_period=Count('status', filter=date_filter & Q(
                status=Order.StatusEnum.RETURNED_FOR_CLARIFICATION.value
            ))
        ).annotate(
            count_all=Count('status', filter=Q(status=Order.StatusEnum.RETURNED_FOR_CLARIFICATION.value))
        ).values('name_param', 'count_period', 'count_all'),
        all=True
    )
    qs = qs.union(
        Order.objects.annotate(
            name_param=Value('Отправлена в обработку')
        ).annotate(
            count_period=Count('status', filter=date_filter & Q(status=Order.StatusEnum.RETURNED_FOR_PROCESS.value))
        ).annotate(
            count_all=Count('status', filter=Q(status=Order.StatusEnum.RETURNED_FOR_PROCESS.value))
        ).values('name_param', 'count_period', 'count_all'),
        all=True
    )
    qs = qs.union(
        Order.objects.values('package_id').annotate(
            name_param=Value('Пакетов')
        ).annotate(
            count_period=Count('package_id', filter=date_filter, distinct=True)
        ).annotate(
            count_all=Count('package_id', distinct=True)
        ).values('name_param', 'count_period', 'count_all'),
        all=True
    )
    qs = qs.union(
        Order.objects.values('author').annotate(
            name_param=Value('Пользователей')
        ).annotate(
            count_period=Count('author', filter=date_filter, distinct=True)
        ).annotate(
            count_all=Count('author', distinct=True)
        ).values('name_param', 'count_period', 'count_all'),
        all=True
    )

    # Допиливаем результат группировкой при помощи groupby из itertools
    for group_name, group_data in groupby(qs, lambda x: x['name_param']):
        count_period = 0
        count_all = 0
        for _g_data in group_data:
            count_period += _g_data['count_period']
            count_all += _g_data['count_all']
        result.append(
            {
                'name_param': group_name,
                'count_period': count_period,
                'count_all': count_all,
            }
        )
    return result


def get_wb(result: List[Dict]) -> openpyxl.Workbook:
    """Полученные данные вставляет в excel файл отчета"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Result'
    ws['A1'] = 'Название'
    ws['B1'] = 'За указанный период'
    ws['C1'] = 'За все время'
    for i, data in enumerate(result, start=2):
        ws[f'A{i}'] = data['name_param']
        ws[f'B{i}'] = data['count_period']
        ws[f'C{i}'] = data['count_all']
    return wb
