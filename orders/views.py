import datetime
from itertools import groupby

import openpyxl
from django.db.models import Count, Q, Value
from django.http import HttpResponse
from django.views.generic.edit import FormView

from orders.forms import IndexForm, ReportForm
from orders.models import Order

field_map = {
    'A': 'number',
    'B': 'state',
    'C': 'agreement',
    'D': 'status',
    'E': 'author',
    'F': 'filename',
    'G': 'created_at',
    'H': 'end_of_work_at',
    'I': 'duration',
    'J': 'fullname',
    'K': 'fullname_after_proccess',
    'L': 'materials_code',
    'M': 'same_materials',
    'N': 'bei',
    'O': 'ntd',
    'P': 'package_id',
}


class IndexFormView(FormView):
    template_name = "index.html"
    form_class = IndexForm
    success_url = "/"

    def post(self, request, *args, **kwargs):
        print(request.FILES)
        excel_file = request.FILES["file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting a particular sheet by name out of many sheets
        worksheet = wb["Data"]
        print(worksheet)

        orders_from_excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        row_number = 0
        for row in worksheet.iter_rows():
            row_number += 1
            if row_number == 1:
                continue
            _order_data = dict()
            for cell in row:
                _order_data[field_map[cell.column_letter]] = (
                    cell.value.strip() if cell.value is not None else cell.value
                )
            _order_data['created_at'] = datetime.datetime.strptime(_order_data['created_at'], '%d.%m.%Y %H:%M:%S')
            _order_data['end_of_work_at'] = datetime.datetime.strptime(
                _order_data['end_of_work_at'], '%d.%m.%Y %H:%M:%S'
            ) if _order_data['end_of_work_at'] else None
            orders_from_excel_data.append(Order(**_order_data))
        Order.objects.bulk_create(orders_from_excel_data)
        return super().post(request, *args, **kwargs)


class ReportView(FormView):
    template_name = "report.html"
    form_class = ReportForm

    def post(self, request, *args, **kwargs):
        result = []
        from_date = datetime.datetime.strptime(
            request.POST['from_date'], '%d.%m.%Y %H:%M:%S'
        ) if request.POST['from_date'] else None
        to_date = datetime.datetime.strptime(
            request.POST['to_date'], '%d.%m.%Y %H:%M:%S'
        ) if request.POST['to_date'] else None
        date_filter = Q(created_at__range=(from_date, to_date))
        qs = Order.objects.annotate(
            name_param=Value('Загруженных заявок')
        ).annotate(
            count_period=Count('id', filter=date_filter)
            # count_period=Order.objects.filter(date_filter).count()
        ).annotate(
            count_all=Count('*')
        ).values('name_param', 'count_period', 'count_all')
        qs = qs.union(
            Order.objects.annotate(
                name_param=Value('Дубли')
            ).annotate(
                count_period=Count('state', filter=date_filter & Q(state__startswith='Дубликат'))
            ).annotate(
                count_all=Count('state', filter=Q(state__startswith='Дубликат'))
            ).values('name_param', 'count_period', 'count_all'),
            all=True
        )
        qs = qs.union(
            Order.objects.annotate(
                name_param=Value('На создание')
            ).annotate(
                count_period=Count('state', filter=date_filter & Q(state__startswith='ДОБАВЛЕНИЕ'))
            ).annotate(
                count_all=Count('state', filter=Q(state__startswith='ДОБАВЛЕНИЕ'))
            ).values('name_param', 'count_period', 'count_all'),
            all=True
        )
        qs = qs.union(
            Order.objects.annotate(
                name_param=Value('На расширение')
            ).annotate(
                count_period=Count('state', filter=date_filter & Q(state__startswith='РАСШИРЕНИЕ'))
            ).annotate(
                count_all=Count('state', filter=Q(state__startswith='РАСШИРЕНИЕ'))
            ).values('name_param', 'count_period', 'count_all'),
            all=True
        )
        qs = qs.union(
            Order.objects.annotate(
                name_param=Value('Обработка завершена')
            ).annotate(
                count_period=Count('status', filter=date_filter & Q(status=Order.StatusEnum.PROCESSED.value))
            ).annotate(
                count_all=Count('status', filter=Q(status=Order.StatusEnum.PROCESSED.value))
            ).values('name_param', 'count_period', 'count_all'),
            all=True
        )
        qs = qs.union(
            Order.objects.annotate(
                name_param=Value('Возвращена на уточнение')
            ).annotate(
                count_period=Count('status', filter=date_filter & Q(
                    status=Order.StatusEnum.RETURNED_FOR_CLARIFICATION.value
                ))
            ).annotate(
                count_all=Count('status', filter=Q(status=Order.StatusEnum.RETURNED_FOR_CLARIFICATION.value))
            ).values('name_param', 'count_period', 'count_all'),
            all=True
        )
        qs = qs.union(
            Order.objects.annotate(
                name_param=Value('Отправлена в обработку')
            ).annotate(
                count_period=Count('status', filter=date_filter & Q(status=Order.StatusEnum.RETURNED_FOR_PROCESS.value))
            ).annotate(
                count_all=Count('status', filter=Q(status=Order.StatusEnum.RETURNED_FOR_PROCESS.value))
            ).values('name_param', 'count_period', 'count_all'),
            all=True
        )
        qs = qs.union(
            Order.objects.values('package_id').annotate(
                name_param=Value('Пакетов')
            ).annotate(
                count_period=Count('package_id', filter=date_filter, distinct=True)
            ).annotate(
                count_all=Count('package_id', distinct=True)
            ).values('name_param', 'count_period', 'count_all'),
            all=True
        )
        qs = qs.union(
            Order.objects.values('author').annotate(
                name_param=Value('Пользователей')
            ).annotate(
                count_period=Count('author', filter=date_filter, distinct=True)
            ).annotate(
                count_all=Count('author', distinct=True)
            ).values('name_param', 'count_period', 'count_all'),
            all=True
        )
        for group_name, group_data in groupby(qs, lambda x: x['name_param']):
            count_period = 0
            count_all = 0
            for _g_data in group_data:
                count_period += _g_data['count_period']
                count_all += _g_data['count_all']
            result.append(
                {
                    'name_param': group_name,
                    'count_period': count_period,
                    'count_all': count_all,
                }
            )
            print(result)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Result'
        ws['A1'] = 'Название'
        ws['B1'] = 'За указанный период'
        ws['C1'] = 'За все время'
        for i, data in enumerate(result, start=2):
            ws[f'A{i}'] = data['name_param']
            ws[f'B{i}'] = data['count_period']
            ws[f'C{i}'] = data['count_all']
        # wb.save('tesssst.xlsx')
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="result.xlsx"'
        wb.save(response)
        # return super().post(request, *args, **kwargs)
        return response
